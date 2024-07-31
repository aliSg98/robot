import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time

class ReportExcel:
    """crear fichero excel"""
    def __init__(self, path_xlsx, logger):
        self.logger = logger
        self.path_xlsx = path_xlsx
        if os.path.exists(self.path_xlsx):
            self.df = pd.read_excel(self.path_xlsx)
            self.logger.setMessage("Archivo Excel encontrado", 'info')
        else:
            # Crear un nuevo DataFrame si el archivo no existe
            new_df = {
                "Name_robot": [],
                "Status_creation": [],
                "Status_pdf": [],
                "Path_pdf": [],
                "Status_final": []
            }
            self.df = pd.DataFrame(new_df)
            self.df.to_excel(self.path_xlsx, index=False, sheet_name='Robots')
            self.logger.setMessage("Archivo Excel no encontrado, nuevo excel creado", 'info')
       

    """Cambiar color del excel, azul claro la primera fila, y luego dependiendo de 
    si es wip, done o fail"""
    def changeColor(self):
        #Cargar el fichero excel con openpyxl
        open_wb = load_workbook(self.path_xlsx)
        res = open_wb.active
        
        #Definir color azul
        color = PatternFill(start_color="ADD8E6",end_color="ADD8E6",fill_type="solid")

        #solo cojo los encabezados de la primera fila
        for col in res[1]:
            col.fill = color
        #guardar excel con el nuevo color
        open_wb.save(self.path_xlsx)

        """Cambiar color segun sea wip -> amarillo, done --> verde, fail --> rojo"""
        color_wip = PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
        color_done = PatternFill(start_color="00FF00",end_color="00FF00",fill_type="solid")
        color_fail = PatternFill(start_color="FF0000",end_color="FF0000",fill_type="solid")

        #columans que me interesa cambiar el color
        c = [2,3,5]
        for columna in res.iter_rows(min_row=2,max_row=res.max_row):
            for col in c:
                #openpyxl usa indices base 1
                celda=columna[col-1]
                #paso a minusculas para que me coja independientemente sea mayus o minus
                cell_value = str(celda.value).strip().lower()
                if cell_value in ['wip']: celda.fill = color_wip
                elif cell_value in ['done']: celda.fill = color_done
                elif cell_value in ['fail']: celda.fill = color_fail
        open_wb.save(self.path_xlsx)
        print("Colores del exel cambiados ")
        self.logger.setMessage("Excel generado con colores",'info')

    """Crear excel vacio"""
    def create_excel(self):
        if os.path.exists(self.path_xlsx):
            print("El archivo Excel ya existe.")
            self.logger.setMessage("El archivo Excel ya existe.", 'info')
        else:
            datos ={
                        "Name_robot": [],
                        "Status_creation": [],
                        "Status_pdf": [],
                        "Path_pdf": [],
                        "Status_final": []
                    }
            self.df = pd.DataFrame(datos)
            os.makedirs(os.path.dirname(self.path_xlsx),exist_ok=True)
            self.df.to_excel(self.path_xlsx, index=False, sheet_name='Robots')
            print("Excel creado sin datos ")
            self.logger.setMessage("Excel creado sin datos ",'info')

    """Añadir datos al excel"""
    def add_data(self, new_data):
        try:
            old_df = pd.read_excel(self.path_xlsx)
            new_df = pd.DataFrame(new_data)
            self.df = pd.concat([old_df, new_df], ignore_index=True) 
            self.logger.setMessage("Datos añadidos al excel ",'info')     
            print("Datos añadidos al excel ")   
        except FileNotFoundError:
            self.df = pd.DataFrame(new_data)
            self.logger.setMessage("Excel no encontrado",'info')
        self.df.to_excel(self.path_xlsx, index=False, sheet_name="Robots")

    """Actualizar una nueva fila o actualizar una fila"""
    def add_update_row_data(self, row_data):
        try:
            column_name = 'Name_robot'
            value = row_data[0]

            # Buscar la fila correspondiente por el valor en la columna
            row_index = self.df.index[self.df[column_name] == value].tolist()

            if row_index:
                # Actualizar la fila existente
                self.df.loc[row_index[0]] = row_data
                self.logger.setMessage(f"Datos del robot: {value} actualizados en el Excel", 'info')
                print(f"Datos del robot: {value} actualizados en el Excel")
            else:
                # Añadir una nueva fila si no se encuentra una fila con el mismo valor
                new_row = pd.DataFrame([row_data], columns=self.df.columns)
                self.df = pd.concat([self.df, new_row], ignore_index=True)
                self.logger.setMessage(f"Nuevo robot: {value} añadido al Excel", 'info')
                print(f"Nuevo robot: {value} añadido al Excel")

            # Guardar el DataFrame actualizado en el archivo Excel
            try:
                self.df.to_excel(self.path_xlsx, index=False, sheet_name='Robots')
                self.logger.setMessage(f"Excel guardado correctamente ", 'info')
            except Exception as e:
                self.logger.setMessage(f"Error al guardar el Excel: {str(e)}", 'error')
                print(f"Error al guardar el Excel: {str(e)}")
        
        except Exception as e:
            self.logger.setMessage(f"Error al actualizar datos", 'error')
            print("Error al actualizar datos")

               

           
            

        